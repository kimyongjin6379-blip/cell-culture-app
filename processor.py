import io
import re
import uuid
import os
import tempfile
from collections import defaultdict

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Constants ────────────────────────────────────────────────────────────────

FEEDING_DAYS = {
    "CHO":       {"Fed-batch": [0, 3, 6], "Batch": []},
    "Hybridoma": {"Fed-batch": [0, 2, 4], "Batch": []},
    "VERO":      {"Fed-batch": [],        "Batch": []},
}

CULTURE_DURATIONS = {
    "CHO":       9,
    "Hybridoma": 6,
    "VERO":     10,
}

SECTION_META = {
    "vcd":        {"label": "Viable Cell Density (VCD)", "unit": "×10⁵ cells/mL",    "chart_type": "line"},
    "ivcd":       {"label": "IVCD",                      "unit": "×10⁵ cell·day/mL", "chart_type": "bar"},
    "viability":  {"label": "Viability",                 "unit": "%",                "chart_type": "line"},
    "titer":      {"label": "Titer",                     "unit": "mg/L",             "chart_type": "bar"},
    "qp":         {"label": "Specific Productivity (Qp)","unit": "pg/cell/day",      "chart_type": "bar"},
    "mu":         {"label": "Specific Growth Rate (μ)",  "unit": "day⁻¹",            "chart_type": "bar"},
}

# Regex patterns
# VCD sample: "251022 D0 SOY-BIO 1 07" or "251022 D3 IMDM 31" or "250907 D6 RICE BRAN #32"
VCD_SAMPLE_RE = re.compile(r"^\s*(\d{6})\s+D(\d+)\s+(.+?)\s+#?(\d{1,3})\s*$", re.I)

# Titer sample: "CHO-29G1 FED D3 SOY-BIO_1" / "PR1 FED D2 IMDM 01" / "CHO-29G1 FED D3 IMDM"
TITER_SAMPLE_RE = re.compile(
    r"^\s*(.+?)\s+(FED|BATCH)(?:-BATCH)?\s+D(\d+)\s+(.+?)\s*$", re.I
)

# ── Cell line / mode detection ───────────────────────────────────────────────

def _detect_from_samples(samples: list, sheet_names: list = None) -> tuple:
    joined = " ".join(str(s).upper() for s in samples if s)
    sn_joined = " ".join((sheet_names or [])).upper()

    # Cell line: prefer sheet names (more reliable than free-text samples)
    if "HYBRIDOMA" in sn_joined or "PR1" in joined or "HYBRIDOMA" in joined:
        cl = "Hybridoma"
    elif "VERO" in sn_joined or "VERO" in joined:
        cl = "VERO"
    elif "CHO" in sn_joined or "CHO" in joined:
        cl = "CHO"
    else:
        cl = "Unknown"

    if "FED" in joined:
        mode = "Fed-batch"
    elif "BATCH" in joined:
        mode = "Batch"
    else:
        mode = "Unknown"
    return cl, mode


# ── Parsing helpers ──────────────────────────────────────────────────────────

def _normalize_name(s: str) -> str:
    """Normalize treatment names for cross-sheet matching (lossy)."""
    s = str(s).upper().strip()
    # Collapse separators
    s = re.sub(r"[\s_\-+]+", "", s)
    # Strip leading zeros inside number runs
    s = re.sub(r"0+(\d)", r"\1", s)
    return s


def _digit_normalize(s: str) -> str:
    """Normalize digit-padding in a treatment name (e.g., 'SOY-BIO 01' == 'SOY-BIO 1')."""
    return re.sub(r"\b0+(\d)", r"\1", str(s).upper().strip())


def _loose_normalize(s: str) -> str:
    """Loose normalization: upper, collapse separators, strip leading zeros, unify '/'↔'+'."""
    s = str(s).upper().replace("/", "+").strip()
    s = re.sub(r"0+(\d)", r"\1", s)        # strip leading zeros
    s = re.sub(r"[\s_\-]+", " ", s).strip()
    return s


def _build_canonical_map(raw_occ: dict, raw_days: dict, d0_treatments: set) -> dict:
    """
    Build raw → canonical treatment name map, merging:
    - Zero-pad / '/'↔'+' variants via loose normalization
    - Trailing feed-volume noise (e.g., 'IMDM 02' → 'IMDM')
    - Missing-suffix typos (e.g., 'PEA' → 'PEA-1') when the shorter form is rare
      and the longer form is frequent.
    """
    # 1) Group by loose_normalize
    loose_groups = defaultdict(list)
    for t in raw_occ:
        loose_groups[_loose_normalize(t)].append(t)

    canon = {}
    for norm, members in loose_groups.items():
        # Prefer a D0 member; else pick the most-frequent
        d0_in = [m for m in members if m in d0_treatments]
        if d0_in:
            rep = max(d0_in, key=lambda x: raw_occ[x])
        else:
            rep = max(members, key=lambda x: raw_occ[x])
        for m in members:
            canon[m] = rep

    # 2) Strip trailing " NN" (feed-volume noise) — if stripped form exists in canon
    working = dict(canon)
    for t, c in list(working.items()):
        m = re.match(r"^(.+?)\s+\d{1,3}$", c)
        if m:
            stripped = m.group(1)
            if stripped in canon and stripped != c:
                canon[t] = canon[stripped]

    # 3) Missing-suffix typos: short rare form is prefix of long frequent form
    # Re-count occurrences after merge
    group_occ = defaultdict(int)
    group_days = defaultdict(set)
    for raw, c in canon.items():
        group_occ[c] += raw_occ[raw]
        group_days[c].update(raw_days[raw])

    reps = list(group_occ.keys())
    # sort by occurrence descending
    reps_sorted = sorted(reps, key=lambda x: -group_occ[x])
    for short in reps:
        if group_occ[short] > 6:   # common enough, skip
            continue
        # candidate longer reps that start with short + '-' or ' '
        candidates = [r for r in reps if r != short
                      and (r.startswith(short + "-") or r.startswith(short + " "))
                      and group_occ[r] >= group_occ[short] * 3]
        if candidates:
            target = max(candidates, key=lambda x: group_occ[x])
            # remap all raws that mapped to short → target
            for raw, c in list(canon.items()):
                if c == short:
                    canon[raw] = target

    return canon


def _parse_vcd_sample(s: str):
    m = VCD_SAMPLE_RE.match(str(s))
    if not m:
        return None
    date, day, treatment, vessel = m.groups()
    return {
        "date": date,
        "day": int(day),
        "treatment": treatment.strip(),
        "vessel": vessel.strip(),
    }


def _parse_titer_sample(s: str, canonical_by_norm: dict):
    """
    Parse titer sample identifiers.
    canonical_by_norm: dict mapping normalized VCD treatment names → canonical VCD treatment
    Returns (canonical_treatment, day, rep_key) or None.
    rep_key is either vessel number (Hybridoma-style) or product variant (CHO _N style).
    """
    m = TITER_SAMPLE_RE.match(str(s))
    if not m:
        return None
    _prefix, _mode, day, tail = m.groups()
    day = int(day)
    tail = tail.strip()

    # CHO-style: "SOY-BIO_1" → treatment="SOY-BIO", rep_key="1"
    # Hybridoma-style: "IMDM 01" → treatment="IMDM", rep_key="01"

    # Try CHO underscore variant first
    und = re.match(r"^(.+?)_(\d+)\s*$", tail)
    if und:
        t_raw, rep_key = und.groups()
        canon = canonical_by_norm.get(_normalize_name(t_raw))
        if canon:
            return canon, day, rep_key.strip()

    # Try trailing vessel number (space-separated)
    vs = re.match(r"^(.+?)\s+(\d{1,3})\s*$", tail)
    if vs:
        t_raw, rep_key = vs.groups()
        canon = canonical_by_norm.get(_normalize_name(t_raw))
        if canon:
            return canon, day, rep_key.strip()

    # Fallback: whole tail as treatment
    canon = canonical_by_norm.get(_normalize_name(tail))
    if canon:
        return canon, day, "1"
    return None


# ── Raw sheet readers ────────────────────────────────────────────────────────

def _read_raw_vcd(xl: pd.ExcelFile):
    """
    Find 'Raw VCD' sheet (Cedex HiRes export). Return dict:
      vcd[treatment][vessel] = {day: {"vcd": float, "viability": float}}
    """
    sheet = None
    for n in xl.sheet_names:
        if "vcd" in n.lower():
            sheet = n; break
    if sheet is None:
        sheet = xl.sheet_names[0]

    df = xl.parse(sheet, header=None, dtype=object)

    # Cedex HiRes puts real headers on row 0
    header_row = 0
    headers = [str(v).strip() if pd.notna(v) else "" for v in df.iloc[header_row].tolist()]

    def _find_col(preds):
        for i, h in enumerate(headers):
            hl = h.lower()
            if any(p(hl) for p in preds):
                return i
        return None

    # Sample identifer (with Cedex typo) / Sample ID fallback
    sample_col = _find_col([
        lambda h: "sample identif" in h,
        lambda h: h == "sample id",
        lambda h: "sample" in h and "id" in h,
    ])
    # Viable Cell Conc. (primary VCD metric) — NOT "Viable Cell Count"
    vcd_col = _find_col([
        lambda h: "viable cell conc" in h,
        lambda h: h == "vcd",
        lambda h: "viable cell density" in h,
    ])
    viab_col = _find_col([
        lambda h: h == "viability" or h.startswith("viability"),
        lambda h: "viab" in h and "%" in h,
    ])

    if sample_col is None or vcd_col is None:
        raise ValueError(f"Could not locate Sample/VCD columns in {sheet}. Headers: {headers}")

    # PASS 1: collect treatment names with occurrence counts across all days
    raw_occ = defaultdict(int)
    raw_days = defaultdict(set)
    d0_treatments = set()
    for r in range(header_row + 1, len(df)):
        sample = df.iloc[r, sample_col]
        if pd.isna(sample):
            continue
        p = _parse_vcd_sample(sample)
        if not p:
            continue
        # Strip trailing " NN" noise and try d0 merge at this stage is skipped
        t = p["treatment"]
        raw_occ[t] += 1
        raw_days[t].add(p["day"])
        if p["day"] == 0:
            d0_treatments.add(t)

    # Build canonical map by fuzzy merging:
    # (a) zero-pad normalization: "SOY-BIO 01" ↔ "SOY-BIO 1"
    # (b) punctuation typo: "CELL BOOST 7A/7B" ↔ "CELL BOOST 7A+7B"
    # (c) missing-suffix typo: "PEA" (rare) ↔ "PEA-1" (frequent)
    # (d) trailing feed-volume noise: "IMDM 02" ↔ "IMDM"
    canon_map = _build_canonical_map(raw_occ, raw_days, d0_treatments)

    vcd: dict = defaultdict(lambda: defaultdict(dict))
    treatment_order: list = []
    all_days: set = set()

    for r in range(header_row + 1, len(df)):
        sample = df.iloc[r, sample_col]
        if pd.isna(sample):
            continue
        parsed = _parse_vcd_sample(sample)
        if not parsed:
            continue
        t = parsed["treatment"]
        v = parsed["vessel"]
        d = parsed["day"]

        # Use canonical map built above
        t = canon_map.get(t, t)

        # Clean vessel number suffix like "32-1"
        v = re.sub(r"-\d+$", "", v)

        all_days.add(d)

        raw_vcd = df.iloc[r, vcd_col] if vcd_col is not None and vcd_col < df.shape[1] else None
        raw_viab = df.iloc[r, viab_col] if viab_col is not None and viab_col < df.shape[1] else None

        try:
            vcd_v = float(raw_vcd) if pd.notna(raw_vcd) else None
        except (ValueError, TypeError):
            vcd_v = None
        try:
            viab_v = float(raw_viab) if pd.notna(raw_viab) else None
        except (ValueError, TypeError):
            viab_v = None

        # Convert VCD to ×10⁵ cells/mL if values look like raw cells/mL (e.g., 1e6)
        if vcd_v is not None and vcd_v > 1e4:
            vcd_v = vcd_v / 1e5

        if t not in treatment_order:
            treatment_order.append(t)

        vcd[t][v][d] = {"vcd": vcd_v, "viability": viab_v}

    return vcd, sorted(all_days), treatment_order


def _read_raw_titer(xl: pd.ExcelFile, canonical_by_norm: dict):
    """
    Find 'Raw Titer' sheet (Cedex Bio export). Return dict:
      titer[treatment][rep_key][day] = value (mg/L)

    For samples like "CHO-29G1 FED D3 SOY-BIO_1" with no trailing vessel number,
    we assign positional rep indices (seq1, seq2, ...) based on row order per (treatment, day).
    For MIGHB/MIGLB rows of same (treatment, rep, day), we SUM (Hybridoma IgG2a).
    """
    sheet = None
    for n in xl.sheet_names:
        if "titer" in n.lower():
            sheet = n; break
    if sheet is None:
        return {}

    df = xl.parse(sheet, header=None, dtype=object)
    header_row = 0
    headers = [str(v).strip() if pd.notna(v) else "" for v in df.iloc[header_row].tolist()]

    def _find_col(preds):
        for i, h in enumerate(headers):
            hl = h.lower()
            if any(p(hl) for p in preds):
                return i
        return None

    sample_col = _find_col([lambda h: h == "sample", lambda h: "sample" in h and "dil" not in h])
    # Cedex Bio result is in "Result" column (mg/L); "Test" indicates assay type
    result_col = _find_col([lambda h: h == "result"])
    test_col = _find_col([lambda h: h == "test"])

    if sample_col is None or result_col is None:
        return {}

    # Raw reads → list of (treatment, rep_key, day, test, value) with positional rep assignment
    # For samples without explicit rep_key, track appearance order per (treatment, day, test)
    titer = defaultdict(lambda: defaultdict(dict))   # [t][rep_key][day] = value
    seq_counter = defaultdict(int)                    # [(t, day, test)] -> next seq num

    for r in range(header_row + 1, len(df)):
        sample = df.iloc[r, sample_col]
        if pd.isna(sample):
            continue
        parsed = _parse_titer_sample(sample, canonical_by_norm)
        raw = df.iloc[r, result_col] if result_col < df.shape[1] else None
        test = str(df.iloc[r, test_col]).strip() if test_col is not None and pd.notna(df.iloc[r, test_col]) else ""

        try:
            v = float(raw) if pd.notna(raw) else None
        except (ValueError, TypeError):
            v = None
        if v is None or parsed is None:
            continue

        treatment, day, rep_key = parsed

        # If rep_key came from parse as literal "1" default (no trailing number in tail),
        # assign positional rep_key per (treatment, day, test)
        tail_has_num = bool(re.search(r"\d", str(sample).strip().split("D" + str(day), 1)[-1]))
        if not tail_has_num:
            seq_counter[(treatment, day, test)] += 1
            rep_key = str(seq_counter[(treatment, day, test)])

        # Handle MIGHB+MIGLB SUM for Hybridoma: same (treatment, rep, day) but different tests
        if day in titer[treatment][rep_key]:
            titer[treatment][rep_key][day] = titer[treatment][rep_key][day] + v
        else:
            titer[treatment][rep_key][day] = v

    return dict(titer)


# ── Computations ─────────────────────────────────────────────────────────────

def _build_vessel_to_rep(vcd: dict) -> dict:
    """For each treatment, sort vessels and assign rep indices."""
    v2r = {}
    for t, vessels in vcd.items():
        sorted_v = sorted(vessels.keys(), key=lambda x: (len(x), x))
        v2r[t] = {v: i + 1 for i, v in enumerate(sorted_v)}
    return v2r


def _compute_ivcd(vcd_by_vessel: dict, all_days: list) -> dict:
    """
    Trapezoidal IVCD per vessel.
    Returns {vessel: {day: ivcd_value}}.
    """
    result = {}
    for vessel, day_data in vcd_by_vessel.items():
        acc = 0.0
        prev_d = None
        prev_v = None
        per_day = {}
        for d in all_days:
            entry = day_data.get(d)
            v = entry["vcd"] if entry else None
            if v is None:
                per_day[d] = None
                continue
            if prev_d is None:
                per_day[d] = 0.0
            else:
                dt = d - prev_d
                acc += (prev_v + v) / 2.0 * dt
                per_day[d] = acc
            prev_d = d
            prev_v = v
        result[vessel] = per_day
    return result


def _compute_mu(vcd_by_vessel: dict, intervals: list) -> dict:
    """
    μ per vessel per interval. intervals = [(start, end), ...].
    Returns {vessel: {(s,e): mu_value}}.
    """
    result = {}
    for vessel, day_data in vcd_by_vessel.items():
        per_int = {}
        for (s, e) in intervals:
            vs = day_data.get(s, {}).get("vcd") if isinstance(day_data.get(s), dict) else None
            ve = day_data.get(e, {}).get("vcd") if isinstance(day_data.get(e), dict) else None
            if vs is not None and ve is not None and vs > 0 and ve > 0 and e > s:
                per_int[(s, e)] = float(np.log(ve / vs) / (e - s))
            else:
                per_int[(s, e)] = None
        result[vessel] = per_int
    return result


def _compute_qp(titer_by_rep: dict, ivcd_by_vessel: dict, v2r: dict, intervals: list) -> dict:
    """
    Qp per rep per interval. Qp = ΔTiter / ΔIVCD (pg/cell/day scale).
    Titer is mg/L, IVCD is ×10⁵ cell·day/mL.
    ΔTiter (mg/L) / ΔIVCD (×10⁵ cell·day/mL) = (mg/L)/(1e5 cell·day/mL)
    = (1e-3 g/L)/(1e5 cell·day / 1e-3 L)  ... simplifying:
    mg/L = 1e-6 g/mL = 1e-6 * 1e12 pg/mL = 1e6 pg/mL
    IVCD ×10⁵ cell·day/mL = 1e5 cell·day/mL
    → (1e6 pg/mL) / (1e5 cell·day/mL) = 10 pg/cell/day
    So multiply by 10.
    Returns {rep_key: {(s,e): qp}}.
    """
    # Build rep → vessel mapping (reverse of v2r)
    rep_to_vessel = {}
    for v, rep in v2r.items():
        rep_to_vessel[str(rep)] = v

    result = {}
    for rep_key, day_titer in titer_by_rep.items():
        # Match rep_key to vessel for IVCD. Try direct, then by rep index.
        vessel = None
        if rep_key in ivcd_by_vessel:
            vessel = rep_key
        elif rep_key in rep_to_vessel:
            vessel = rep_to_vessel[rep_key]
        else:
            # Try stripping leading zeros
            stripped = rep_key.lstrip("0") or "0"
            if stripped in rep_to_vessel:
                vessel = rep_to_vessel[stripped]

        per_int = {}
        for (s, e) in intervals:
            ts = day_titer.get(s)
            te = day_titer.get(e)
            ivcd_data = ivcd_by_vessel.get(vessel, {}) if vessel else {}
            is_ = ivcd_data.get(s)
            ie_ = ivcd_data.get(e)
            if (ts is not None and te is not None
                    and is_ is not None and ie_ is not None
                    and (ie_ - is_) > 0):
                per_int[(s, e)] = float((te - ts) * 10.0 / (ie_ - is_))
            else:
                per_int[(s, e)] = None
        result[rep_key] = per_int
    return result


def _stats_from_rows(rows: list) -> tuple:
    """rows: list of lists (per rep). Return (mean_list, std_list)."""
    if not rows:
        return [], []
    n = max(len(r) for r in rows)
    means, stds = [], []
    for i in range(n):
        vals = [r[i] for r in rows if i < len(r) and r[i] is not None]
        if vals:
            means.append(round(float(np.mean(vals)), 4))
            stds.append(round(float(np.std(vals, ddof=1)) if len(vals) > 1 else 0.0, 4))
        else:
            means.append(None)
            stds.append(None)
    return means, stds


# ── Public API ───────────────────────────────────────────────────────────────

def process_file(file_bytes: bytes, basal_media: str = "", feed_media: str = "") -> dict:
    xl = pd.ExcelFile(io.BytesIO(file_bytes))

    vcd, all_days, treatment_order = _read_raw_vcd(xl)
    canonical_by_norm = {_normalize_name(t): t for t in treatment_order}

    # Detect cell line / mode from samples
    samples = []
    for n in xl.sheet_names:
        try:
            d = xl.parse(n, header=None, dtype=object, nrows=200)
            for col in range(min(3, d.shape[1])):
                samples.extend(d.iloc[:, col].dropna().astype(str).tolist()[:50])
        except Exception:
            continue
    cell_line, culture_mode = _detect_from_samples(samples, xl.sheet_names)

    titer = _read_raw_titer(xl, canonical_by_norm)

    v2r = _build_vessel_to_rep(vcd)

    # Compute IVCD per vessel
    ivcd_all = {}
    for t, vessels in vcd.items():
        ivcd_all[t] = _compute_ivcd(vessels, all_days)

    # μ / Qp intervals: match researcher's 정리-sheet convention.
    # For Fed-batch with feeding days: use only consecutive feeding-day intervals
    # (exclude post-last-feed stretch, where cells enter stationary/death phase).
    feeding = FEEDING_DAYS.get(cell_line, {}).get(culture_mode, [])

    # Collect titer days
    titer_days = set()
    for t in titer.values():
        for rep in t.values():
            titer_days.update(rep.keys())
    titer_days = sorted(titer_days)

    if feeding and len(feeding) >= 2:
        # Fed-batch: feeding-day-based intervals only
        mu_intervals = [(feeding[i], feeding[i + 1]) for i in range(len(feeding) - 1)]
        # Qp: same intervals, then filtered later if all-None (e.g. no D0 titer)
        qp_intervals = mu_intervals[:]
    else:
        # Batch or VERO: consecutive measurement days
        mu_intervals = [(all_days[i], all_days[i + 1]) for i in range(len(all_days) - 1)] \
                        if len(all_days) >= 2 else []
        qp_intervals = [(titer_days[i], titer_days[i + 1]) for i in range(len(titer_days) - 1)] \
                        if len(titer_days) >= 2 else []

    # Compute μ & Qp
    mu_all = {t: _compute_mu(vessels, mu_intervals) for t, vessels in vcd.items()}
    qp_all = {}
    for t in treatment_order:
        if t in titer:
            qp_all[t] = _compute_qp(titer[t], ivcd_all.get(t, {}), v2r.get(t, {}), qp_intervals)

    # Build chart-friendly sections
    result_sections = _build_sections(
        treatment_order, vcd, ivcd_all, titer, mu_all, qp_all,
        all_days, titer_days, mu_intervals, qp_intervals, v2r,
    )

    # Build Excel
    file_id = str(uuid.uuid4())
    out_path = os.path.join(tempfile.gettempdir(), f"{file_id}.xlsx")
    _build_jeongri_excel(
        result_sections, cell_line, culture_mode,
        basal_media, feed_media, out_path,
        all_days, titer_days, mu_intervals, qp_intervals,
        treatment_order,
    )

    return {
        "cell_line": cell_line,
        "culture_mode": culture_mode,
        "feeding_days": feeding,
        "sections": result_sections,
        "title": f"{cell_line} {culture_mode}",
        "file_id": file_id,
    }


def _build_sections(treatment_order, vcd, ivcd_all, titer, mu_all, qp_all,
                    all_days, titer_days, mu_intervals, qp_intervals, v2r):
    sections = {}

    # VCD
    vcd_treats = {}
    for t in treatment_order:
        sorted_vessels = sorted(vcd[t].keys(), key=lambda x: (len(x), x))
        rows = []
        for v in sorted_vessels:
            rows.append([vcd[t][v].get(d, {}).get("vcd") if vcd[t][v].get(d) else None for d in all_days])
        means, stds = _stats_from_rows(rows)
        vcd_treats[t] = {"mean": means, "std": stds, "replicates": rows}
    sections["vcd"] = {**SECTION_META["vcd"], "days": all_days,
                       "x_labels": [f"D{d}" for d in all_days], "treatments": vcd_treats}

    # Viability
    via_treats = {}
    for t in treatment_order:
        sorted_vessels = sorted(vcd[t].keys(), key=lambda x: (len(x), x))
        rows = []
        for v in sorted_vessels:
            rows.append([vcd[t][v].get(d, {}).get("viability") if vcd[t][v].get(d) else None for d in all_days])
        means, stds = _stats_from_rows(rows)
        via_treats[t] = {"mean": means, "std": stds, "replicates": rows}
    sections["viability"] = {**SECTION_META["viability"], "days": all_days,
                             "x_labels": [f"D{d}" for d in all_days], "treatments": via_treats}

    # IVCD
    ivcd_treats = {}
    for t in treatment_order:
        sorted_vessels = sorted(ivcd_all.get(t, {}).keys(), key=lambda x: (len(x), x))
        rows = [[ivcd_all[t][v].get(d) for d in all_days] for v in sorted_vessels]
        means, stds = _stats_from_rows(rows)
        ivcd_treats[t] = {"mean": means, "std": stds, "replicates": rows}
    sections["ivcd"] = {**SECTION_META["ivcd"], "days": all_days,
                        "x_labels": [f"D{d}" for d in all_days], "treatments": ivcd_treats}

    # Titer
    if titer and titer_days:
        titer_treats = {}
        for t in treatment_order:
            if t not in titer:
                continue
            rep_keys = sorted(titer[t].keys(), key=lambda x: (len(x), x))
            rows = [[titer[t][rk].get(d) for d in titer_days] for rk in rep_keys]
            means, stds = _stats_from_rows(rows)
            titer_treats[t] = {"mean": means, "std": stds, "replicates": rows}
        sections["titer"] = {**SECTION_META["titer"], "days": titer_days,
                             "x_labels": [f"D{d}" for d in titer_days], "treatments": titer_treats}

    # μ — filter intervals where ALL treatments have no data
    mu_treats_full = {}
    for t in treatment_order:
        sorted_vessels = sorted(mu_all.get(t, {}).keys(), key=lambda x: (len(x), x))
        rows = [[mu_all[t][v].get((s, e)) for (s, e) in mu_intervals] for v in sorted_vessels]
        means, stds = _stats_from_rows(rows)
        mu_treats_full[t] = {"mean": means, "std": stds, "replicates": rows}
    # Determine which intervals have at least one non-None value
    keep_idx = [i for i in range(len(mu_intervals))
                if any(mu_treats_full[t]["mean"][i] is not None for t in mu_treats_full
                       if i < len(mu_treats_full[t]["mean"]))]
    kept_mu_intervals = [mu_intervals[i] for i in keep_idx]
    mu_labels = [f"D{s}-D{e}" for (s, e) in kept_mu_intervals]
    mu_days_only = [e for (_s, e) in kept_mu_intervals]
    mu_treats = {}
    for t, stat in mu_treats_full.items():
        mu_treats[t] = {
            "mean": [stat["mean"][i] for i in keep_idx],
            "std":  [stat["std"][i]  for i in keep_idx],
            "replicates": [[rep[i] for i in keep_idx] for rep in stat["replicates"]],
        }
    sections["mu"] = {**SECTION_META["mu"], "days": mu_days_only,
                      "x_labels": mu_labels, "treatments": mu_treats}

    # Qp — same filtering
    if qp_all and qp_intervals:
        qp_treats_full = {}
        for t in treatment_order:
            if t not in qp_all:
                continue
            rep_keys = sorted(qp_all[t].keys(), key=lambda x: (len(x), x))
            rows = [[qp_all[t][rk].get((s, e)) for (s, e) in qp_intervals] for rk in rep_keys]
            means, stds = _stats_from_rows(rows)
            qp_treats_full[t] = {"mean": means, "std": stds, "replicates": rows}
        keep_idx = [i for i in range(len(qp_intervals))
                    if any(qp_treats_full[t]["mean"][i] is not None for t in qp_treats_full
                           if i < len(qp_treats_full[t]["mean"]))]
        kept_qp = [qp_intervals[i] for i in keep_idx]
        qp_labels = [f"D{s}-D{e}" for (s, e) in kept_qp]
        qp_days_only = [e for (_s, e) in kept_qp]
        qp_treats = {}
        for t, stat in qp_treats_full.items():
            qp_treats[t] = {
                "mean": [stat["mean"][i] for i in keep_idx],
                "std":  [stat["std"][i]  for i in keep_idx],
                "replicates": [[rep[i] for i in keep_idx] for rep in stat["replicates"]],
            }
        sections["qp"] = {**SECTION_META["qp"], "days": qp_days_only,
                          "x_labels": qp_labels, "treatments": qp_treats}

    return sections


# ── Excel output (정리-style, single sheet) ──────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_SUBHDR_FILL = PatternFill("solid", fgColor="2E75B6")
_ALT_FILL    = PatternFill("solid", fgColor="D6E4F0")
_WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
_META_FILL   = PatternFill("solid", fgColor="FFF2CC")
_THIN = Side(style="thin", color="AAAAAA")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _hdr(ws, r, c, v, fill=None, color="FFFFFF", bold=True, size=11):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(bold=bold, color=color, name="Arial", size=size)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _BORDER
    if fill:
        cell.fill = fill
    return cell


def _data(ws, r, c, v, fill=None, fmt="0.000"):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name="Arial", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _BORDER
    cell.number_format = fmt
    if fill:
        cell.fill = fill
    return cell


def _build_jeongri_excel(sections, cell_line, culture_mode, basal_media, feed_media,
                         path, all_days, titer_days, mu_intervals, qp_intervals,
                         treatment_order):
    wb = Workbook()
    ws = wb.active
    ws.title = "정리"

    max_days = max(len(all_days), len(titer_days), len(mu_intervals), len(qp_intervals))
    total_cols = 3 + max_days

    # Metadata header
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    _hdr(ws, 1, 1, f"{cell_line} {culture_mode} Cell Culture Summary", fill=_HEADER_FILL, size=14)
    ws.row_dimensions[1].height = 26

    meta_row = 2
    ws.cell(row=meta_row, column=1, value="Cell Line").font = Font(bold=True)
    ws.cell(row=meta_row, column=2, value=cell_line)
    ws.cell(row=meta_row, column=3, value="Mode").font = Font(bold=True)
    ws.cell(row=meta_row, column=4, value=culture_mode)
    meta_row += 1
    ws.cell(row=meta_row, column=1, value="Basal Media").font = Font(bold=True)
    ws.merge_cells(start_row=meta_row, start_column=2, end_row=meta_row, end_column=total_cols)
    ws.cell(row=meta_row, column=2, value=basal_media or "-")
    meta_row += 1
    ws.cell(row=meta_row, column=1, value="Feed Media").font = Font(bold=True)
    ws.merge_cells(start_row=meta_row, start_column=2, end_row=meta_row, end_column=total_cols)
    feed_display = feed_media or ", ".join(treatment_order)
    ws.cell(row=meta_row, column=2, value=feed_display)
    meta_row += 2

    row = meta_row

    sec_order = ["vcd", "ivcd", "viability", "titer", "mu", "qp"]
    for sec_key in sec_order:
        if sec_key not in sections:
            continue
        sec = sections[sec_key]
        days = sec["days"]
        x_labels = sec["x_labels"]
        treatments = sec["treatments"]

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
        _hdr(ws, row, 1, f"{sec['label']} ({sec['unit']})", fill=_HEADER_FILL, size=12)
        ws.row_dimensions[row].height = 22
        row += 1

        _hdr(ws, row, 1, "Treatment", fill=_SUBHDR_FILL)
        _hdr(ws, row, 2, "Stat", fill=_SUBHDR_FILL)
        for j, lbl in enumerate(x_labels):
            _hdr(ws, row, 3 + j, lbl, fill=_SUBHDR_FILL)
        row += 1

        for t_idx, (name, stat) in enumerate(treatments.items()):
            fill = _ALT_FILL if t_idx % 2 == 0 else _WHITE_FILL

            # Individual reps
            for rep_i, rep in enumerate(stat["replicates"]):
                ws.cell(row=row, column=1, value=name if rep_i == 0 else "").font = Font(bold=True, name="Arial", size=10)
                ws.cell(row=row, column=1).fill = fill
                ws.cell(row=row, column=1).border = _BORDER
                _hdr(ws, row, 2, f"rep{rep_i + 1}", fill=fill, color="000000", bold=False)
                for j, v in enumerate(rep):
                    _data(ws, row, 3 + j, v, fill=fill)
                row += 1

            # Mean
            ws.cell(row=row, column=1, value="").fill = fill
            ws.cell(row=row, column=1).border = _BORDER
            _hdr(ws, row, 2, "Mean", fill=fill, color="000000", bold=True)
            for j, v in enumerate(stat["mean"]):
                _data(ws, row, 3 + j, v, fill=fill)
            row += 1

            # SD
            ws.cell(row=row, column=1, value="").fill = fill
            ws.cell(row=row, column=1).border = _BORDER
            _hdr(ws, row, 2, "SD", fill=fill, color="000000", bold=True)
            for j, v in enumerate(stat["std"]):
                _data(ws, row, 3 + j, v, fill=fill)
            row += 1

        row += 1  # spacer

    # Column widths
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 8
    for j in range(max_days):
        ws.column_dimensions[get_column_letter(3 + j)].width = 11

    wb.save(path)
